# %%
import numpy as np


def calculate_neck_distance(mouse1_neck, mouse2_neck):
    """
    Calculate the distance between the necks of two mice from the origin (0,0)
    and the distance between the neck of mouse 1 and mouse 2.

    Parameters:
    mouse1_neck (tuple): (x1, y1) coordinates of the neck of mouse 1
    mouse2_neck (tuple): (x2, y2) coordinates of the neck of mouse 2

    Returns:
    tuple: (distance from origin to neck of mouse 1,
            distance from origin to neck of mouse 2,
            distance between necks of mouse 1 and mouse 2)
    """

    x1, y1 = mouse1_neck
    x2, y2 = mouse2_neck

    # Distance from the origin (0,0) to the neck of mouse 1
    d1 = np.sqrt(x1**2 + y1**2)

    # Distance from the origin (0,0) to the neck of mouse 2
    d2 = np.sqrt(x2**2 + y2**2)

    # Distance between the necks of mouse 1 and mouse 2
    d_neck = np.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)

    return d1, d2, d_neck


# Example usage:
mouse1_neck = (2, 3)  # Replace with the actual coordinates of mouse 1's neck
mouse2_neck = (5, 7)  # Replace with the actual coordinates of mouse 2's neck

d1, d2, d_neck = calculate_neck_distance(mouse1_neck, mouse2_neck)

print(f"Distance from origin to mouse 1's neck: {d1}")
print(f"Distance from origin to mouse 2's neck: {d2}")
print(f"Distance between the necks of mouse 1 and mouse 2: {d_neck}")

# %%
import numpy as np


def berechne_hypotenuse(punkt1, punkt2):
    """
    Berechnet die Hypotenuse (Abstand) zwischen zwei Punkten im 2D-Raum.

    Parameter:
    punkt1 (tuple): Koordinaten von Punkt 1 als (x1, y1)
    punkt2 (tuple): Koordinaten von Punkt 2 als (x2, y2)

    Rückgabe:
    float: Die Hypotenuse (Abstand) zwischen den beiden Punkten
    """

    x1, y1 = punkt1
    x2, y2 = punkt2

    # Berechnung der Hypotenuse zwischen den zwei Punkten
    hypotenuse = np.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)

    return hypotenuse


# Beispiel:
punkt1 = (3, 4)  # Beispielkoordinaten für Punkt 1
punkt2 = (7, 1)  # Beispielkoordinaten für Punkt 2

hypotenuse = berechne_hypotenuse(punkt1, punkt2)
print(
    f"Die Hypotenuse zwischen den Punkten {punkt1} und {punkt2} beträgt: {hypotenuse}"
)

# %%
import pandas as pd

# Pfad zur CSV-Datei (Beispiel)
file_path = "C:/Users/DeinBenutzername/Ordnerstruktur/deinedatei.csv"

# Einlesen der CSV-Datei
df = pd.read_csv(file_path)

# Anzeige der ersten paar Zeilen der Datei
print(df.head())

# %%
import pandas as pd
from openpyxl import load_workbook


# Funktion zum Eintragen von Daten in eine Excel-Datei
def schreibe_ergebnisse_in_excel(ergebnisse, excel_pfad, spalten_name):
    """
    Schreibt Ergebnisse nacheinander in eine Excel-Tabelle.

    Parameter:
    ergebnisse (list): Liste der Ergebnisse für jede Maus.
    excel_pfad (str): Pfad zur Excel-Datei.
    spalten_name (str): Name der Spalte, in die die Ergebnisse geschrieben werden.
    """
    try:
        # Lade die bestehende Excel-Datei
        writer = pd.ExcelWriter(excel_pfad, engine="openpyxl")
        writer.book = load_workbook(excel_pfad)
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

        # Erstelle ein DataFrame für die Ergebnisse
        df_ergebnisse = pd.DataFrame({spalten_name: ergebnisse})

        # Schreibe die Ergebnisse in die nächste freie Spalte
        df_ergebnisse.to_excel(
            writer,
            sheet_name="Tabelle1",
            index=False,
            startcol=writer.sheets["Tabelle1"].max_column,
        )

        # Speichere die Excel-Datei
        writer.save()
        print(f"Ergebnisse erfolgreich in {spalten_name} eingetragen.")

    except FileNotFoundError:
        # Falls die Datei nicht existiert, erstelle eine neue Excel-Datei
        df_ergebnisse = pd.DataFrame({spalten_name: ergebnisse})
        df_ergebnisse.to_excel(excel_pfad, index=False)
        print(f"Neue Datei erstellt und Ergebnisse in {spalten_name} eingetragen.")


# Beispielergebnisse für Maus 1 und Maus 2 (Hypotenusen)
ergebnisse_maus1 = [2.5, 3.1, 4.2, 5.6]  # Hypotenusen für Maus 1
ergebnisse_maus2 = [3.5, 2.9, 4.8, 6.0]  # Hypotenusen für Maus 2

# Pfad zur Excel-Datei
excel_pfad = "C:/Users/DeinBenutzername/Ergebnisse.xlsx"

# Ergebnisse für Maus 1 eintragen
schreibe_ergebnisse_in_excel(ergebnisse_maus1, excel_pfad, "Maus 1")

# Ergebnisse für Maus 2 eintragen
schreibe_ergebnisse_in_excel(ergebnisse_maus2, excel_pfad, "Maus 2")
